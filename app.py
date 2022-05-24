import json
import re
import time

import requests
import random

from openpyxl import load_workbook

ta = []
tb = []
tc = []

file = '3.xlsx'
max_a = load_workbook(file).active.max_row
for a in range(0, max_a + 1):
    ta.append(0)
    tb.append(0)
    tc.append(0)


def pj(p):
    p = str(p)
    l = p[3:7]

    n = p.replace(l, "****")
    return n


def writeexcle(t):
    wb = load_workbook(file)
    sheet = wb.active
    max_row = sheet.max_row - t

    row_max = 'a' + str(max_row)
    print(sheet)

    if max_row > 0:
        a = sheet[row_max].value
        print(a)
        return a
    else:
        print("空了")
        return None


def rt():
    t = random.randint(1, 6)
    print('随即停止', t)
    time.sleep(t)


def getuid(phone):
    print("获取uid")
    rt()
    url = 'https://www.chaojijishi.com/api/v3/thirdPartyCheckRedirect'

    header1 = {
        'user-agent': 'Mozilla/5.0 (Linux; Android 11; M2006J10C Build/RP1A.200720.011; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/98.0.4758.101 Mobile Safari/537.36'
    }
    header2 = {
        'user-agent': 'Mozilla/5.0 (Linux; U; Android 8.1.0; zh-cn; BLA-AL00 Build/HUAWEIBLA-AL00) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/57.0.2987.132 MQQBrowser/8.9 Mobile Safari/537.36'
    }
    header8 = {
        'user-agent': 'Mozilla/5.0 (Linux; Android 8.1; PAR-AL00 Build/HUAWEIPAR-AL00; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/57.0.2987.132 MQQBrowser/6.2 TBS/044304 Mobile Safari/537.36 MicroMessenger/6.7.3.1360(0x26070333) NetType/WIFI Language/zh_CN Process/tools'}
    header7 = {
        'user-agent': 'Mozilla/5.0 (Linux; Android 8.1.0; ALP-AL00 Build/HUAWEIALP-AL00; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/63.0.3239.83 Mobile Safari/537.36 T7/10.13 baiduboxapp/10.13.0.11 (Baidu; P1 8.1.0)'}
    header6 = {
        'user-agent': 'Mozilla/5.0 (Linux; U; Android 8.0.0; zh-CN; MHA-AL00 Build/HUAWEIMHA-AL00) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/57.0.2987.108 UCBrowser/12.1.4.994 Mobile Safari/537.36'}
    header5 = {
        'user-agent': 'Mozilla/5.0 (Linux; Android 8.0; MHA-AL00 Build/HUAWEIMHA-AL00; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/57.0.2987.132 MQQBrowser/6.2 TBS/044304 Mobile Safari/537.36 MicroMessenger/6.7.3.1360(0x26070333) NetType/NON_NETWORK Language/zh_CN Process/tools'}
    header4 = {
        'user-agent': 'Mozilla/5.0 (Linux; U; Android 4.1.2; zh-cn; HUAWEI MT1-U06 Build/HuaweiMT1-U06) AppleWebKit/534.30 (KHTML, like Gecko) Version/4.0 Mobile Safari/534.30 baiduboxapp/042_2.7.3_diordna_8021_027/IEWAUH_61_2.1.4_60U-1TM+IEWAUH/7300001a/91E050E40679F078E51FD06CD5BF0A43%7C544176010472968/1'}
    header3 = {
        'user-agent': 'Mozilla/5.0 (Linux; Android 8.0; MHA-AL00 Build/HUAWEIMHA-AL00; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/57.0.2987.132 MQQBrowser/6.2 TBS/044304 Mobile Safari/537.36 MicroMessenger/6.7.3.1360(0x26070333) NetType/4G Language/zh_CN Process/tools'
    }

    data = {
        "mobile": str(phone),
        "password": "123456",
        "type": "zl_jiudou"
    }
    p = random.randint(1, 8)
    r = None
    if p == 1:
        r = requests.post(url=url, headers=header1, data=data)
    if p == 2:
        r = requests.post(url=url, headers=header2, data=data)
    if p == 3:
        r = requests.post(url=url, headers=header3, data=data)
    if p == 4:
        r = requests.post(url=url, headers=header4, data=data)
    if p == 5:
        r = requests.post(url=url, headers=header5, data=data)
    if p == 6:
        r = requests.post(url=url, headers=header6, data=data)
    if p == 7:
        r = requests.post(url=url, headers=header7, data=data)
    if p == 8:
        r = requests.post(url=url, headers=header8, data=data)
    k = r.text
    # print(k)
    rs = re.findall('user_id\":(.+?),', k)
    print("userid", rs[0])
    # rt()
    return rs[0]


def gettk(phone, id1):
    print("获取tk")
    # 登录ck
    url1 = 'https://v1.kykykb.cn/third_login?partner_id=xiaomi&platform_code=xiaomi&third_username=' + pj(str(
        phone)) + '&third_user_id=' + str(
        id1) + '&device_token=E65EDB645421EA65F709E6975338483AF55F2F76&device_type=xiaomi&version=1.2.2&platform=android&third_type=jiudou&timestamp=' + str(
        int(round(time.time() * 1000)))

    header2 = {
        'accept-encoding': 'gzip',
        'user-agent': 'okhttp/3.14.9'
    }

    r = requests.post(url=url1, headers=header2)
    x = r.text
    # print(x)
    s = re.findall('token\":\"(.+?)\"', x)
    # print(s)
    # print(s[0])
    return s[0]


def qd(header):
    print("签到")
    rt()
    url2 = 'https://v1.kykykb.cn/task/sign_in'
    r = requests.get(url=url2, headers=header)
    res = r.content.decode('utf-8')
    res = json.loads(res)
    # print(res)
    res = str(res)
    s = re.findall('code\': (\d.+?\d),', res)
    return s


def dz(header):
    print("点赞")
    rt()
    url4 = 'https://v1.kykykb.cn/forum/section/hot?partner_id=xiaomi&platform_code=xiaomi&page=1&version=1.2.2&platform=android&timestamp=' + str(
        int(round(time.time() * 1000)))
    r = requests.post(url=url4, headers=header)
    li = r.text
    lis = re.findall('id\":(\d+\.?\d),\"title', li)
    print("当前长度")
    print(len(lis))
    m = 0
    for l in lis:
        rt()
        url3 = 'https://v1.kykykb.cn/forum/topic/agree?topic_id=' + l
        r = requests.post(url=url3, headers=header)
        res = r.content.decode('utf-8')
        res = json.loads(res)
        print(res)
        m += 1
        if m >= 5:
            break
    url6 = 'https://v1.kykykb.cn/task/receive?partner_id=xiaomi&platform_code=xiaomi&id=3&version=1.2.2&platform=android&timestamp=' + str(
        int(round(time.time() * 1000)))
    r = requests.get(url=url6, headers=header)
    res = r.content.decode('utf-8')
    res = json.loads(res)
    # print(res)
    res = str(res)
    s = re.findall('code\': (\d.+?\d),', res)
    return s


def gettask(header):
    print("任务")
    rt()
    ti = int(round(time.time() * 1000))
    url = 'https://v1.kykykb.cn/task/list?partner_id=xiaomi&platform_code=xiaomi&version=1.2.2&platform=android&timestamp=' + str(
        ti)
    r = requests.get(url=url, headers=header)
    res = r.content.decode('utf-8')
    res = json.loads(res)
    # print(r.content)
    print(res)

    a = str(res)
    s = re.findall('state_desc\': \'(.+?)\'', a)
    a3 = s[0]
    a4 = s[2]
    a5 = s[4]
    # print(s)
    print(a3, a4, a5)
    return a3, a4, a5


def fb(header):
    print("发表")
    rt()
    tim = str(int(round(time.time() * 1000)))
    url5 = 'https://v1.kykykb.cn/forum/topic/edit?partner_id=xiaomi&platform_code=xiaomi&category_id=687&resources=%5B%7B%22type%22%3A%22text%22%2C%22content%22%3A%222%22%7D%5D&title=1&version=1.2.2&platform=android&timestamp=' + tim
    url6 = 'https://v1.kykykb.cn/forum/topic/edit?partner_id=xiaomi&platform_code=xiaomi&category_id=3188&resources=%5B%7B%22type%22%3A%22text%22%2C%22content%22%3A%22%E6%9D%A5%E4%BA%86%22%7D%5D&title=&version=1.2.2&platform=android&timestamp=' + tim
    url7 = 'https://v1.kykykb.cn/forum/topic/edit?partner_id=xiaomi&platform_code=xiaomi&category_id=689&resources=%5B%7B%22type%22%3A%22text%22%2C%22content%22%3A%22%E5%BF%AB%E4%B9%90%22%7D%5D&title=&version=1.2.2&platform=android&timestamp=' + tim
    url8 = 'https://v1.kykykb.cn/forum/topic/edit?partner_id=xiaomi&platform_code=xiaomi&category_id=6735&resources=%5B%7B%22type%22%3A%22text%22%2C%22content%22%3A%22%E7%BE%8E%E5%A5%BD%22%7D%5D&title=&version=1.2.2&platform=android&timestamp=1' + tim
    ks = random.randint(5, 8)
    if ks == 5:
        requests.post(url=url5, headers=header)
    if ks == 6:
        requests.post(url=url6, headers=header)
    if ks == 7:
        requests.post(url=url7, headers=header)
    if ks == 8:
        requests.post(url=url8, headers=header)

    # res = r.content.decode('utf-8')
    # res = json.loads(res)
    # print(res)
    url7 = 'https://v1.kykykb.cn/task/receive?partner_id=xiaomi&platform_code=xiaomi&id=7&version=1.2.2&platform=android&timestamp=' + str(
        int(round(time.time() * 1000)))
    r = requests.get(url=url7, headers=header)
    res = r.content.decode('utf-8')
    res = json.loads(res)
    # print(res)
    res = str(res)
    s = re.findall('code\': (\d.+?\d),', res)
    return s


def lg(header, t, id1):
    x = 0
    y = t
    t = max_a - t
    a = gettask(header)
    if a[0] != "完成":
        l1 = qd(header)
        if l1 == '1103':
            header = setheader(gettk(writeexcle(y), id1))
            l1 = qd(header)
        if l1 == '200':
            x += 1
            ta[t] = 1
    else:
        x += 1
        ta[t] = 1
    if a[1] != "完成":
        l2 = dz(header)
        if l2 == '1103':
            header = setheader(gettk(writeexcle(y), id1))
            l2 = dz(header)
        if l2 == '200':
            x += 1
            tb[t] = 1
    else:
        x += 1
        tb[t] = 1
    if a[2] != "完成":
        l3 = fb(header)
        if l3 == '1103':
            header = setheader(gettk(writeexcle(y), id1))
            l3 = fb(header)
        if l3 == '200':
            x += 1
            tc[t] = 1
    else:
        x += 1
        tc[t] = 1
    return x


def setheader(tk):
    header = {
        'Host': 'v1.kykykb.cn',
        'token': tk.replace('\\', ''),
        'accept-encoding': 'gzip',
        'user-agent': 'okhttp/3.14.9'

    }
    return header


def ww(k):
    b = 0
    for iq in k:
        if iq == 0:
            print(max_a - b)
        b += 1


errs = []
errph = []
t = 0
jt = 0
while t < max_a:
    try:
        phone = writeexcle(t)
        id1 = getuid(phone)
        max_b = lg(setheader(gettk(phone, id1)), t, id1)
        err = 0
        while max_b != 3:
            max_b = lg(setheader(gettk(phone, id1)), t, id1)
            err += 1
            if err > 3:
                print("失败一个")
                errph.append(phone)
                max_b = 3
        t += 1
        jt = 0
    except Exception as es:
        print("异常结束", es)
        print("失败序号", max_a - t)
        print("当前t值", t)
        errs.append(t)
        jt += 1
        t += 1
        if jt >= 3:
            print("网络异常，已退出")
            break

print('失败数：')
for ers in errph:
    print(ers)

print("任务一完成")
a1 = 0
for i in ta:
    a1 += i
print(a1)
print("任务二完成")
a2 = 0
for i in tb:
    a2 += i
print(a2)
print("任务三完成")
a3 = 0
for i in tc:
    a3 += i
print(a3)

# print("各任务失败序号")
# print("任务一")
# ww(ta)
# print("任务二")
# ww(tb)
# print("任务三")
# ww(tc)

print("登录异常的")
for ia in errs:
    print(ia)

print("当前t值", t)
